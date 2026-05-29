# auditoria_engine.py
# Motor de auditoria FreteScan Pro - versão corrigida para o padrão ATUA x GW
# Foco: leitura por blocos/células extraídas do PDF, não por linha única.

import io
import json
import os
import re
import tempfile
import unicodedata
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
import pdfplumber


CENTAVOS = Decimal("0.01")

RE_CTE_NUM = re.compile(r"^\d{1,6}$")
RE_DATA_ATUA = re.compile(r"^\d{2}/\d{2}/\d{2}\s+\d{2}:\d{2}$")
RE_DATA_GW = re.compile(r"^\d{2}/\d{2}/\d{4}$")
RE_PLACA = re.compile(r"^[A-Z]{3}\d[A-Z0-9]\d{2}$|^[A-Z]{3}\d{4}$")
RE_MONEY_BR = re.compile(r"^-?(?:\d{1,3}(?:\.\d{3})+,\d{2}|\d+,\d{2}|\d+\.\d{2})$")
MONEY_RE = re.compile(r"-?(?:\d{1,3}(?:\.\d{3})+,\d{2}|\d+,\d{2}|\d+\.\d{2})")
RE_PESO_TON = re.compile(r"^\d{1,3},\d{3}$")
RE_PERCENT = re.compile(r"^-?\d{1,3},\d{2}%$")
MAX_PDF_PAGE_COUNT = 300
MAX_HISTORY_ENTRIES = 100

LAST_PARSE_DEBUG = {
    "ATUA": {},
    "GW": {},
}


def parse_money_br(value) -> Optional[Decimal]:
    """
    Converte moeda brasileira para Decimal corretamente.

    Exemplos:
    23.919,00 -> 23919.00
    24.839,65 -> 24839.65
    1.817,84  -> 1817.84
    0,00      -> 0.00
    """
    if value is None:
        return None

    if isinstance(value, Decimal):
        return value.quantize(CENTAVOS, rounding=ROUND_HALF_UP)

    text = str(value).strip()
    text = text.replace("R$", "").replace(" ", "")

    if text in ["", "-", "None", "nan"]:
        return None

    # BR: ponto = milhar, vírgula = decimal
    if "," in text:
        text = text.replace(".", "")
        text = text.replace(",", ".")
    else:
        text = text.replace(",", "")

    try:
        return Decimal(text).quantize(CENTAVOS, rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return None


def decimal_to_float(value):
    if value is None:
        return None
    return float(value)


def format_money_br(value) -> str:
    if value is None:
        return "-"

    if not isinstance(value, Decimal):
        value = Decimal(str(value))

    value = value.quantize(CENTAVOS, rounding=ROUND_HALF_UP)
    sinal = "-" if value < 0 else ""
    value_abs = abs(value)
    s = f"{value_abs:.2f}"
    inteiro, centavos = s.split(".")
    partes = []

    while len(inteiro) > 3:
        partes.insert(0, inteiro[-3:])
        inteiro = inteiro[:-3]

    partes.insert(0, inteiro)
    return f"{sinal}R$ {'.'.join(partes)},{centavos}"


def normalizar_cte(value) -> Optional[str]:
    if value is None:
        return None

    text = str(value).strip()

    if not re.fullmatch(r"\d{1,6}", text):
        return None

    numero = int(text)

    if numero < 1 or numero > 999999:
        return None

    return str(numero)


def _selecionar_valores_gw(valores: List[Decimal]) -> tuple[Optional[Decimal], Optional[Decimal]]:
    if len(valores) >= 10:
        return valores[1], valores[9]
    nao_zero = [valor for valor in valores if valor != Decimal("0.00")]
    if len(nao_zero) >= 2:
        return nao_zero[0], nao_zero[-1]
    if len(valores) >= 2:
        return valores[0], valores[-1]
    return None, None


def _extrair_linhas_pdfplumber(caminho_pdf):
    linhas = []
    with pdfplumber.open(str(caminho_pdf)) as pdf:
        total_paginas = len(pdf.pages)
        if total_paginas == 0:
            raise ValueError("O PDF enviado não possui páginas legíveis.")
        if total_paginas > MAX_PDF_PAGE_COUNT:
            raise ValueError(
                f"O PDF possui {total_paginas} páginas e excede o limite de {MAX_PDF_PAGE_COUNT} páginas."
            )
        for page_num, page in enumerate(pdf.pages, start=1):
            texto = page.extract_text() or ""
            for raw in texto.splitlines():
                t = raw.strip()
                if t:
                    linhas.append((page_num, t))
    return linhas


def extrair_linhas_pymupdf(caminho_pdf):
    import fitz

    linhas = []
    with fitz.open(str(caminho_pdf)) as doc:
        total_paginas = len(doc)
        if total_paginas == 0:
            raise ValueError("O PDF enviado nao possui paginas legiveis.")
        if total_paginas > MAX_PDF_PAGE_COUNT:
            raise ValueError(
                f"O PDF possui {total_paginas} paginas e excede o limite de {MAX_PDF_PAGE_COUNT} paginas."
            )
        for page_index, page in enumerate(doc, start=1):
            texto = page.get_text("text") or ""
            for linha in texto.splitlines():
                linha = " ".join(str(linha).split())
                if linha:
                    linhas.append((page_index, linha))
    return linhas


def _normalizar_texto_busca(texto: str) -> str:
    texto = unicodedata.normalize("NFKD", texto or "")
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    return texto.lower()


def _detectar_cabecalho_pr(linhas_pdf) -> Optional[str]:
    texto = _normalizar_texto_busca("\n".join(linha for _, linha in linhas_pdf[:80]))
    if "relatorio detalhado do ctrc" in texto or ("detalhado do ctrc" in texto and "ct" in texto):
        return "ATUA PR"
    if "analise de cte/nfs com impostos" in texto or "cte/nfs com impostos" in texto:
        return "GW PR"
    return None


def _detectar_formato_pr(linhas_pdf) -> Optional[str]:
    formato = _detectar_cabecalho_pr(linhas_pdf)
    if formato:
        return formato
    if any(RE_ATUA_PR_INICIO.match(" ".join(str(linha).split())) for _, linha in linhas_pdf[:200]):
        return "ATUA PR"
    if any(RE_GW_PR_INICIO.match(" ".join(str(linha).split())) for _, linha in linhas_pdf[:200]):
        return "GW PR"
    return None


def _linha_parece_cte(tipo: str, linha: str) -> bool:
    linha = " ".join(str(linha).split())
    if tipo == "ATUA":
        return bool(RE_ATUA_PR_INICIO.match(linha) or RE_ATUA_LINHA.match(linha))
    if tipo == "GW":
        return bool(RE_GW_PR_INICIO.match(linha) or RE_GW_LINHA.match(linha))
    return False


def _registrar_debug_parser(tipo: str, formato: str, metodo: str, registros, linhas_pdf):
    ordenados = [registros[cte] for cte in sorted(registros.keys(), key=lambda x: int(x))[:10]]
    paginas_lidas = sorted({page_num for page_num, _ in linhas_pdf})
    linhas_cte = [linha for _, linha in linhas_pdf if _linha_parece_cte(tipo, linha)]
    LAST_PARSE_DEBUG[tipo] = {
        "layout_detectado": tipo,
        "formato_detectado": formato,
        "metodo_usado": metodo,
        "paginas_lidas": len(paginas_lidas),
        "linhas_lidas": len(linhas_pdf),
        "linhas_parecem_cte": len(linhas_cte),
        "primeiras_10_linhas_parecem_cte": linhas_cte[:10],
        "quantidade_ctes": len(registros),
        "primeiros_10_ctes": [r["cte"] for r in ordenados],
        "primeiros_10_valores_empresa": [r["empresa"] for r in ordenados],
        "primeiros_10_valores_motorista": [r["motorista"] for r in ordenados],
        "primeiras_10_linhas_reais_extraidas": [linha for _, linha in linhas_pdf[:10]],
    }


# Regex para linha de CTE do ATUA:  "1752 CT ..."
RE_ATUA_LINHA = re.compile(r"^\s*(\d{1,6})\s+CT\b")
# Regex para linha de CTE do GW: "001752 01/04/2026 ..."
RE_GW_LINHA = re.compile(r"^\s*(\d{4,6})\s+\d{2}/\d{2}/\d{4}\b")
RE_ATUA_PR_INICIO = re.compile(r"^\s*(\d{1,6})\s+CT\s+\d{2}/\d{2}/\d{2}\s+\d{2}:\d{2}\s+")
RE_GW_PR_INICIO = re.compile(r"^\s*(\d{6})\s+\d{2}/\d{2}/\d{4}\s+")
RE_PLACA_PR = re.compile(r"\b(?:[A-Z]{3}[0-9][A-Z0-9][0-9]{2}|[A-Z]{3}[0-9]{4})\b")
RE_NUM_BR_PR = re.compile(r"-?\d{1,3}(?:\.\d{3})*,\d{2,3}")
RE_MONEY_GW_PR = re.compile(r"-?\d{1,3}(?:\.\d{3})*,\d{2}(?!%)")
RE_MARGEM_GW_PR = re.compile(r"-?\d{1,3},\d{2}%")

ATUA_HEADER_LINES = {
    "Numero",
    "T.",
    "Emissao Hora",
    "Filial",
    "Agencia",
    "Remetente",
    "Destinatario",
    "Pagador",
    "Placa",
    "Peso (Ton)",
    "Frete Empr.",
    "Frete Mot.",
    "Adto. Empr.",
    "Adto. Mot.",
}

GW_HEADER_LINES = {
    "CTe/NFS",
    "Emissao",
    "Remetente / Origem",
    "Destinatario / Destino",
    "Tipo Frete",
    "Peso / Kg",
    "Valor frete",
    "ICMS/ISS (%)",
    "Frete tab.",
    "PIS",
    "COFINS",
    "IR",
    "CSSL",
    "Vl Carreteiro Liquido",
}


def _extrair_atua_linha_unica(linhas_pdf) -> Dict[str, Dict[str, Any]]:
    registros = {}
    for page_num, linha in linhas_pdf:
        m = RE_ATUA_LINHA.match(linha)
        if not m:
            continue
        cte = normalizar_cte(m.group(1))
        if not cte:
            continue

        placas = list(RE_PLACA_PR.finditer(linha))
        if placas:
            resto = linha[placas[-1].end():]
            todos = RE_NUM_BR_PR.findall(resto)
        else:
            todos = MONEY_RE.findall(linha)

        if len(todos) < 3:
            continue

        peso = parse_money_br(todos[0]) if placas else None
        empresa_a = parse_money_br(todos[1])
        motorista_a = parse_money_br(todos[2])

        if empresa_a is None or motorista_a is None:
            continue

        registros[cte] = {
            "cte": cte,
            "placa": placas[-1].group(0) if placas else None,
            "peso": peso,
            "frete_empresa": empresa_a,
            "frete_motorista": motorista_a,
            "empresa": empresa_a,
            "motorista": motorista_a,
            "pagina": page_num,
            "margem": None,
            "raw": linha
        }
    return registros


def _ignorar_linha_atua_multilinha(linha: str) -> bool:
    return (
        linha in ATUA_HEADER_LINES
        or linha.startswith("ATUA - ")
        or linha.startswith("Pagina:")
        or linha.startswith("Relatorio Detalhado do CTRC")
        or linha.startswith("Ambiente")
    )


def _finalizar_bloco_atua(registros, bloco):
    if not bloco:
        return

    valores = [
        parse_money_br(linha)
        for linha in bloco["linhas"]
        if RE_MONEY_BR.fullmatch(linha)
    ]
    valores = [valor for valor in valores if valor is not None]

    if len(valores) < 2:
        return

    registros[bloco["cte"]] = {
        "cte": bloco["cte"],
        "placa": None,
        "peso": None,
        "frete_empresa": valores[0],
        "frete_motorista": valores[1],
        "empresa": valores[0],
        "motorista": valores[1],
        "pagina": bloco["pagina"],
        "margem": None,
        "raw": " | ".join(bloco["linhas"]),
    }


def _extrair_atua_multilinha(linhas_pdf) -> Dict[str, Dict[str, Any]]:
    registros = {}
    bloco_atual = None

    for page_num, linha in linhas_pdf:
        if _ignorar_linha_atua_multilinha(linha):
            continue

        cte = normalizar_cte(linha) if RE_CTE_NUM.fullmatch(linha) else None
        if cte:
            _finalizar_bloco_atua(registros, bloco_atual)
            bloco_atual = {"cte": cte, "pagina": page_num, "linhas": []}
            continue

        if bloco_atual is not None:
            bloco_atual["linhas"].append(linha)

    _finalizar_bloco_atua(registros, bloco_atual)
    return registros


def _extrair_atua_pr_multilinha(linhas_pdf) -> Dict[str, Dict[str, Any]]:
    return _extrair_atua_multilinha(linhas_pdf)


def extrair_atua_pr(caminho_pdf, linhas_pdf=None) -> Dict[str, Dict[str, Any]]:
    linhas_pdf = linhas_pdf if linhas_pdf is not None else _extrair_linhas_pdfplumber(caminho_pdf)
    registros = {}

    for page_num, linha in linhas_pdf:
        linha = " ".join(str(linha).split())
        m = RE_ATUA_PR_INICIO.match(linha)
        if not m:
            continue

        cte = normalizar_cte(m.group(1))
        if not cte:
            continue

        placas = list(RE_PLACA_PR.finditer(linha))
        if not placas:
            continue

        resto = linha[placas[-1].end():]
        nums = RE_NUM_BR_PR.findall(resto)
        if len(nums) < 3:
            continue

        peso = parse_money_br(nums[0])
        empresa_a = parse_money_br(nums[1])
        motorista_a = parse_money_br(nums[2])
        if empresa_a is None or motorista_a is None:
            continue

        registros[cte] = {
            "cte": cte,
            "placa": placas[-1].group(0),
            "peso": peso,
            "frete_empresa": empresa_a,
            "frete_motorista": motorista_a,
            "empresa": empresa_a,
            "motorista": motorista_a,
            "pagina": page_num,
            "margem": None,
            "raw": linha,
        }

    return registros


def extrair_atua_por_blocos(caminho_pdf) -> Dict[str, Dict[str, Any]]:
    # Reutiliza a mesma extração textual para evitar varrer o PDF duas vezes.
    linhas_pdf = _extrair_linhas_pdfplumber(caminho_pdf)
    formato_cabecalho = _detectar_cabecalho_pr(linhas_pdf)
    if formato_cabecalho == "ATUA PR":
        registros = extrair_atua_pr(caminho_pdf, linhas_pdf)
        _registrar_debug_parser("ATUA", formato_cabecalho, "pdfplumber", registros, linhas_pdf)
        return registros

    registros = _extrair_atua_linha_unica(linhas_pdf)
    if registros:
        _registrar_debug_parser("ATUA", "ATUA legado", "parser atual", registros, linhas_pdf)
        return registros
    registros = _extrair_atua_multilinha(linhas_pdf)
    if registros:
        _registrar_debug_parser("ATUA", "ATUA legado", "parser atual", registros, linhas_pdf)
        return registros

    formato = _detectar_formato_pr(linhas_pdf)
    if formato == "ATUA PR":
        registros = extrair_atua_pr(caminho_pdf, linhas_pdf)
        _registrar_debug_parser("ATUA", formato, "pdfplumber", registros, linhas_pdf)
        return registros

    _registrar_debug_parser("ATUA", formato or "nao detectado", "pdfplumber", registros, linhas_pdf)
    return registros


def _extrair_gw_linha_unica(linhas_pdf) -> Dict[str, Dict[str, Any]]:
    registros = {}
    for page_num, linha in linhas_pdf:
        m = RE_GW_LINHA.match(linha)
        if not m:
            continue
        cte = normalizar_cte(m.group(1))
        if not cte:
            continue

        _, _, depois_combinado = linha.partition("Combinado")
        origem_valores = depois_combinado if depois_combinado else linha
        valores = [parse_money_br(item) for item in MONEY_RE.findall(origem_valores)]
        valores = [valor for valor in valores if valor is not None]

        if len(valores) < 2:
            continue

        # Empresa B usa "Valor frete" do GW. Não usar "Frete tab." porque pode vir líquido/descontado por impostos.
        empresa_b, motorista_b = _selecionar_valores_gw(valores)

        if empresa_b is None or motorista_b is None:
            continue

        registros[cte] = {
            "cte": cte,
            "peso": valores[0] if len(valores) > 0 else None,
            "valor_frete": empresa_b,
            "frete_tab": valores[4] if len(valores) > 4 else None,
            "vl_carreteiro_liquido": motorista_b,
            "empresa": empresa_b,
            "motorista": motorista_b,
            "pagina": page_num,
            "margem": None,
            "raw": linha
        }
    return registros


def _ignorar_linha_gw_multilinha(linha: str) -> bool:
    return (
        linha in GW_HEADER_LINES
        or linha.startswith("GW - ")
        or linha.startswith("Pagina:")
        or linha.startswith("Analise de CTe/NFS")
        or linha.startswith("Usuario:")
        or linha.startswith("FILIAL :")
    )


def _finalizar_bloco_gw(registros, bloco):
    if not bloco:
        return

    cte = None
    pagina_cte = bloco["pagina"]
    valores_antes_cte = []

    for page_num, linha in bloco["linhas"]:
        if RE_CTE_NUM.fullmatch(linha):
            cte = normalizar_cte(linha)
            pagina_cte = page_num
            break

        if RE_MONEY_BR.fullmatch(linha):
            valor = parse_money_br(linha)
            if valor is not None:
                valores_antes_cte.append(valor)

    empresa_b, motorista_b = _selecionar_valores_gw(valores_antes_cte)
    if not cte or empresa_b is None or motorista_b is None:
        return

    registros[cte] = {
        "cte": cte,
        # Empresa B usa "Valor frete" do GW. Não usar "Frete tab." porque pode vir líquido/descontado por impostos.
        "peso": valores_antes_cte[0] if len(valores_antes_cte) > 0 else None,
        "valor_frete": empresa_b,
        "frete_tab": valores_antes_cte[4] if len(valores_antes_cte) > 4 else None,
        "vl_carreteiro_liquido": motorista_b,
        "empresa": empresa_b,
        "motorista": motorista_b,
        "pagina": pagina_cte,
        "margem": None,
        "raw": " | ".join(linha for _, linha in bloco["linhas"]),
    }


def _extrair_gw_multilinha(linhas_pdf) -> Dict[str, Dict[str, Any]]:
    registros = {}
    bloco_atual = None

    for page_num, linha in linhas_pdf:
        if _ignorar_linha_gw_multilinha(linha):
            continue

        if RE_DATA_GW.fullmatch(linha):
            _finalizar_bloco_gw(registros, bloco_atual)
            bloco_atual = {"pagina": page_num, "linhas": [(page_num, linha)]}
            continue

        if bloco_atual is not None:
            bloco_atual["linhas"].append((page_num, linha))

    _finalizar_bloco_gw(registros, bloco_atual)
    return registros


def _extrair_gw_pr_multilinha(linhas_pdf) -> Dict[str, Dict[str, Any]]:
    registros = {}
    bloco_atual = []

    def finalizar(bloco):
        if not bloco:
            return
        cte = None
        pagina = bloco[0][0]
        valores = []
        for page_num, linha in bloco:
            linha = " ".join(str(linha).split())
            if cte is None and RE_CTE_NUM.fullmatch(linha):
                cte = normalizar_cte(linha)
                pagina = page_num
            for item in RE_MONEY_GW_PR.findall(linha):
                valor = parse_money_br(item)
                if valor is not None:
                    valores.append(valor)

        if not cte or len(valores) < 2:
            return

        empresa_b = valores[0]
        motorista_b = valores[9] if len(valores) >= 10 else valores[2] if len(valores) >= 3 else valores[-1]
        registros[cte] = {
            "cte": cte,
            "peso": None,
            "valor_frete": empresa_b,
            "frete_tab": valores[4] if len(valores) > 4 else None,
            "vl_carreteiro_liquido": motorista_b,
            "empresa": empresa_b,
            "motorista": motorista_b,
            "pagina": pagina,
            "margem": None,
            "raw": " | ".join(linha for _, linha in bloco),
        }

    for page_num, linha in linhas_pdf:
        linha_normalizada = " ".join(str(linha).split())
        if RE_DATA_GW.fullmatch(linha_normalizada):
            finalizar(bloco_atual)
            bloco_atual = [(page_num, linha_normalizada)]
            continue
        if bloco_atual:
            bloco_atual.append((page_num, linha_normalizada))

    finalizar(bloco_atual)
    return registros


def extrair_gw_pr(caminho_pdf, linhas_pdf=None) -> Dict[str, Dict[str, Any]]:
    linhas_pdf = linhas_pdf if linhas_pdf is not None else _extrair_linhas_pdfplumber(caminho_pdf)
    registros = {}

    for idx, (page_num, linha) in enumerate(linhas_pdf):
        linha = " ".join(str(linha).split())
        m = RE_GW_PR_INICIO.match(linha)
        if not m:
            continue

        cte = normalizar_cte(m.group(1))
        if not cte:
            continue

        _, _, depois_combinado = linha.partition("Combinado")
        if not depois_combinado:
            continue

        valores_txt = RE_MONEY_GW_PR.findall(depois_combinado)
        if len(valores_txt) < 10:
            continue

        valores = [parse_money_br(item) for item in valores_txt]
        if any(valor is None for valor in valores[:10]):
            continue

        peso = valores[0]
        valor_frete = valores[1]
        icms_iss = valores[2]
        percentual_icms = valores[3]
        frete_tab = valores[4]
        pis = valores[5]
        cofins = valores[6]
        ir = valores[7]
        cssl = valores[8]
        vl_carreteiro_liquido = valores[9]
        diferenca = valores[10] if len(valores) > 10 else None

        margem = None
        pct = RE_MARGEM_GW_PR.findall(depois_combinado)
        if pct:
            margem = pct[-1]
        else:
            for j in range(idx + 1, min(idx + 4, len(linhas_pdf))):
                linha_pct = " ".join(str(linhas_pdf[j][1]).split())
                pct = RE_MARGEM_GW_PR.findall(linha_pct)
                if pct:
                    margem = pct[-1]
                    break

        registros[cte] = {
            "cte": cte,
            "peso": peso,
            "valor_frete": valor_frete,
            "icms_iss": icms_iss,
            "percentual_icms": percentual_icms,
            "frete_tab": frete_tab,
            "pis": pis,
            "cofins": cofins,
            "ir": ir,
            "cssl": cssl,
            "vl_carreteiro_liquido": vl_carreteiro_liquido,
            "diferenca_gw": diferenca,
            "empresa": valor_frete,
            "motorista": vl_carreteiro_liquido,
            "pagina": page_num,
            "margem": margem,
            "raw": linha,
        }

    return registros


def extrair_gw_por_blocos(caminho_pdf) -> Dict[str, Dict[str, Any]]:
    # Reutiliza a mesma extração textual para evitar varrer o PDF duas vezes.
    linhas_pdf = _extrair_linhas_pdfplumber(caminho_pdf)
    formato_cabecalho = _detectar_cabecalho_pr(linhas_pdf)
    if formato_cabecalho == "GW PR":
        registros = extrair_gw_pr(caminho_pdf, linhas_pdf)
        _registrar_debug_parser("GW", formato_cabecalho, "pdfplumber", registros, linhas_pdf)
        return registros

    registros = _extrair_gw_linha_unica(linhas_pdf)
    if registros:
        _registrar_debug_parser("GW", "GW legado", "parser atual", registros, linhas_pdf)
        return registros
    registros = _extrair_gw_multilinha(linhas_pdf)
    if registros:
        _registrar_debug_parser("GW", "GW legado", "parser atual", registros, linhas_pdf)
        return registros

    formato = _detectar_formato_pr(linhas_pdf)
    if formato == "GW PR":
        registros = extrair_gw_pr(caminho_pdf, linhas_pdf)
        _registrar_debug_parser("GW", formato, "pdfplumber", registros, linhas_pdf)
        return registros

    _registrar_debug_parser("GW", formato or "nao detectado", "pdfplumber", registros, linhas_pdf)
    return registros


def ler_atua(caminho_pdf):
    registros = extrair_atua_por_blocos(caminho_pdf)

    if not registros:
        raise ValueError("Falha na leitura do ATUA. Nenhum CTE válido com valores foi encontrado.")

    return registros


def _bytes_para_tmp(pdf_bytes: bytes) -> str:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(pdf_bytes)
        return tmp.name

def parse_atua(pdf_bytes: bytes) -> tuple[pd.DataFrame, list[str]]:
    tmp = _bytes_para_tmp(pdf_bytes)
    try:
        registros = extrair_atua_por_blocos(tmp)
    finally:
        os.unlink(tmp)
    linhas = [{"CTE": r["cte"], "EmpresaA": r["empresa"], "MotoristaA": r["motorista"]} for r in registros.values()]
    return pd.DataFrame(linhas), []

def parse_gw(pdf_bytes: bytes) -> tuple[pd.DataFrame, list[str]]:
    tmp = _bytes_para_tmp(pdf_bytes)
    try:
        registros = extrair_gw_por_blocos(tmp)
    finally:
        os.unlink(tmp)
    linhas = [{"CTE": r["cte"], "EmpresaB": r["empresa"], "MotoristaB": r["motorista"]} for r in registros.values()]
    return pd.DataFrame(linhas), []



def ler_gw(caminho_pdf):
    registros = extrair_gw_por_blocos(caminho_pdf)

    if not registros:
        raise ValueError("Falha na leitura do GW. Nenhum CTE válido com valores foi encontrado.")

    return registros


def calcular_status(existe_a, existe_b, dif_empresa, dif_motorista, tolerancia: Decimal):
    if existe_a and not existe_b:
        return "Faltante no GW"

    if existe_b and not existe_a:
        return "Faltante no ATUA"

    abs_emp = abs(dif_empresa)
    abs_mot = abs(dif_motorista)

    if abs_emp == Decimal("0.00") and abs_mot == Decimal("0.00"):
        return "OK"

    if abs_emp <= tolerancia and abs_mot <= tolerancia:
        return "OK por arredondamento"

    return "Divergente"


def auditar(caminho_atua, caminho_gw, tolerancia=Decimal("0.50")):
    if not isinstance(tolerancia, Decimal):
        tolerancia = Decimal(str(tolerancia)).quantize(CENTAVOS, rounding=ROUND_HALF_UP)

    registros_a = ler_atua(caminho_atua)
    registros_b = ler_gw(caminho_gw)

    validar_integridade_basica(registros_a, registros_b)

    todos_ctes = sorted(
        set(registros_a.keys()) | set(registros_b.keys()),
        key=lambda x: int(x)
    )

    linhas = []

    for cte in todos_ctes:
        a = registros_a.get(cte)
        b = registros_b.get(cte)

        existe_a = a is not None
        existe_b = b is not None

        empresa_a = a["empresa"] if a else None
        motorista_a = a["motorista"] if a else None
        empresa_b = b["empresa"] if b else None
        motorista_b = b["motorista"] if b else None

        # Para cálculo de faltante, usa 0 apenas no cálculo, mas mantém campo exibido como None.
        calc_empresa_a = empresa_a if empresa_a is not None else Decimal("0.00")
        calc_empresa_b = empresa_b if empresa_b is not None else Decimal("0.00")
        calc_motorista_a = motorista_a if motorista_a is not None else Decimal("0.00")
        calc_motorista_b = motorista_b if motorista_b is not None else Decimal("0.00")

        dif_empresa = (calc_empresa_a - calc_empresa_b).quantize(CENTAVOS, rounding=ROUND_HALF_UP)
        dif_motorista = (calc_motorista_a - calc_motorista_b).quantize(CENTAVOS, rounding=ROUND_HALF_UP)
        maior_diferenca = max(abs(dif_empresa), abs(dif_motorista))

        status = calcular_status(existe_a, existe_b, dif_empresa, dif_motorista, tolerancia)

        linhas.append({
            "CTE": cte,
            "Status": status,
            "Empresa A": empresa_a,
            "Empresa B": empresa_b,
            "Motorista A": motorista_a,
            "Motorista B": motorista_b,
            "Dif Empresa": dif_empresa,
            "Dif Motorista": dif_motorista,
            "Maior Diferença": maior_diferenca,
            "Margem B": b.get("margem") if b else None,
            "Página A": a.get("pagina") if a else None,
            "Página B": b.get("pagina") if b else None,
        })

    resumo = gerar_resumo(linhas, registros_a, registros_b, tolerancia)

    return {
        "resumo": resumo,
        "linhas": linhas,
        "registros_a": registros_a,
        "registros_b": registros_b,
        "debug": gerar_debug(registros_a, registros_b),
    }


def gerar_resumo(linhas, registros_a, registros_b, tolerancia):
    total = len(linhas)

    def count(status):
        return sum(1 for x in linhas if x["Status"] == status)

    dif_empresa_total = sum((x["Dif Empresa"] for x in linhas), Decimal("0.00")).quantize(CENTAVOS)
    dif_motorista_total = sum((x["Dif Motorista"] for x in linhas), Decimal("0.00")).quantize(CENTAVOS)

    impacto_confirmado = sum(
        (x["Maior Diferença"] for x in linhas if x["Status"] in ["Divergente", "Faltante no GW"]),
        Decimal("0.00")
    ).quantize(CENTAVOS)
    impacto_potencial_total = sum(
        (x["Maior Diferença"] for x in linhas if x["Status"] in ["Divergente", "Faltante no ATUA", "Faltante no GW"]),
        Decimal("0.00")
    ).quantize(CENTAVOS)

    return {
        "tolerancia": tolerancia,
        "ctes_atua": len(registros_a),
        "ctes_gw": len(registros_b),
        "total_analisado": total,
        "cruzados": len(set(registros_a.keys()) & set(registros_b.keys())),
        "ok": count("OK"),
        "ok_arredondamento": count("OK por arredondamento"),
        "divergentes": count("Divergente"),
        "faltante_a": count("Faltante no ATUA"),
        "faltante_b": count("Faltante no GW"),
        "volume_faltante_atua": count("Faltante no ATUA"),
        "dif_empresa_total": dif_empresa_total,
        "dif_motorista_total": dif_motorista_total,
        "impacto_absoluto": impacto_confirmado,
        "impacto_critico_confirmado": impacto_confirmado,
        "impacto_potencial_total": impacto_potencial_total,
        "dif_total_empresa": dif_empresa_total,    # aliases for frontend
        "dif_total_motorista": dif_motorista_total # aliases for frontend
    }


def gerar_resumo_df(df: pd.DataFrame) -> dict:
    total = len(df)
    ok = len(df[df["Status"] == "OK"])
    ok_r = len(df[df["Status"] == "OK por arredondamento"])
    div = len(df[df["Status"] == "Divergente"])
    fa = len(df[df["Status"] == "Faltante no ATUA"])
    fb = len(df[df["Status"] == "Faltante no GW"])
    crit = df[df["Status"].isin(["Divergente", "Faltante no GW"])]
    potencial = df[df["Status"].isin(["Divergente", "Faltante no ATUA", "Faltante no GW"])]
    dif_empresa_total = df["Dif Empresa"].fillna(0).sum()
    dif_motorista_total = df["Dif Motorista"].fillna(0).sum()
    impacto_abs = crit["Maior Diferença"].fillna(0).sum()
    impacto_potencial_total = potencial["Maior Diferença"].fillna(0).sum()

    return {
        "total": int(total),
        "ok": int(ok),
        "ok_arredondamento": int(ok_r),
        "divergentes": int(div),
        "faltantes_a": int(fa),
        "faltantes_b": int(fb),
        "volume_faltante_atua": int(fa),
        "dif_total_empresa": float(round(dif_empresa_total, 2)),
        "dif_total_motorista": float(round(dif_motorista_total, 2)),
        "impacto_absoluto": float(round(impacto_abs, 2)),
        "impacto_critico_confirmado": float(round(impacto_abs, 2)),
        "impacto_potencial_total": float(round(impacto_potencial_total, 2)),
    }

def validar_integridade_basica(registros_a, registros_b):
    erros = []

    if len(registros_a) == 0:
        erros.append("ATUA retornou 0 CTEs.")

    if len(registros_b) == 0:
        erros.append("GW retornou 0 CTEs.")

    if registros_b:
        absurdos_b = sum(1 for r in registros_b.values() if r["motorista"] > Decimal("200000.00"))
        if absurdos_b > len(registros_b) * Decimal("0.05"):
            erros.append("GW com valores absurdos no Motorista B. Conversão monetária provavelmente errada.")

    inter = set(registros_a.keys()) & set(registros_b.keys())

    if len(inter) == 0:
        erros.append("Nenhum CTE cruzado entre ATUA e GW.")

    if erros:
        raise ValueError("Falha na validação de integridade: " + " | ".join(erros))


def gerar_debug(registros_a, registros_b):
    def top(registros):
        saida = []
        for cte in sorted(registros.keys(), key=lambda x: int(x))[:10]:
            r = registros[cte]
            saida.append({
                "CTE": cte,
                "Empresa": r["empresa"],
                "Motorista": r["motorista"],
                "Página": r.get("pagina")
            })
        return saida

    return {
        "ATUA - layout detectado": LAST_PARSE_DEBUG.get("ATUA", {}).get("layout_detectado"),
        "ATUA - formato detectado": LAST_PARSE_DEBUG.get("ATUA", {}).get("formato_detectado"),
        "ATUA - metodo usado": LAST_PARSE_DEBUG.get("ATUA", {}).get("metodo_usado"),
        "ATUA - paginas lidas": LAST_PARSE_DEBUG.get("ATUA", {}).get("paginas_lidas"),
        "ATUA - linhas lidas": LAST_PARSE_DEBUG.get("ATUA", {}).get("linhas_lidas"),
        "ATUA - linhas que parecem CTE": LAST_PARSE_DEBUG.get("ATUA", {}).get("linhas_parecem_cte"),
        "ATUA - primeiras 10 linhas que parecem CTE": LAST_PARSE_DEBUG.get("ATUA", {}).get("primeiras_10_linhas_parecem_cte", []),
        "ATUA - quantidade de CTEs": LAST_PARSE_DEBUG.get("ATUA", {}).get("quantidade_ctes", len(registros_a)),
        "ATUA - primeiros 10 CTEs": LAST_PARSE_DEBUG.get("ATUA", {}).get("primeiros_10_ctes", []),
        "ATUA - primeiros 10 valores empresa": LAST_PARSE_DEBUG.get("ATUA", {}).get("primeiros_10_valores_empresa", []),
        "ATUA - primeiros 10 valores motorista": LAST_PARSE_DEBUG.get("ATUA", {}).get("primeiros_10_valores_motorista", []),
        "ATUA - primeiras 10 linhas reais extraidas": LAST_PARSE_DEBUG.get("ATUA", {}).get("primeiras_10_linhas_reais_extraidas", []),
        "GW - layout detectado": LAST_PARSE_DEBUG.get("GW", {}).get("layout_detectado"),
        "GW - formato detectado": LAST_PARSE_DEBUG.get("GW", {}).get("formato_detectado"),
        "GW - metodo usado": LAST_PARSE_DEBUG.get("GW", {}).get("metodo_usado"),
        "GW - paginas lidas": LAST_PARSE_DEBUG.get("GW", {}).get("paginas_lidas"),
        "GW - linhas lidas": LAST_PARSE_DEBUG.get("GW", {}).get("linhas_lidas"),
        "GW - linhas que parecem CTE": LAST_PARSE_DEBUG.get("GW", {}).get("linhas_parecem_cte"),
        "GW - primeiras 10 linhas que parecem CTE": LAST_PARSE_DEBUG.get("GW", {}).get("primeiras_10_linhas_parecem_cte", []),
        "GW - quantidade de CTEs": LAST_PARSE_DEBUG.get("GW", {}).get("quantidade_ctes", len(registros_b)),
        "GW - primeiros 10 CTEs": LAST_PARSE_DEBUG.get("GW", {}).get("primeiros_10_ctes", []),
        "GW - primeiros 10 valores empresa": LAST_PARSE_DEBUG.get("GW", {}).get("primeiros_10_valores_empresa", []),
        "GW - primeiros 10 valores motorista": LAST_PARSE_DEBUG.get("GW", {}).get("primeiros_10_valores_motorista", []),
        "GW - primeiras 10 linhas reais extraidas": LAST_PARSE_DEBUG.get("GW", {}).get("primeiras_10_linhas_reais_extraidas", []),
        "ATUA - Top 10": top(registros_a),
        "GW - Top 10": top(registros_b),
    }


def linhas_para_dataframe(linhas) -> pd.DataFrame:
    rows = []

    for x in linhas:
        rows.append({
            "CTE": x["CTE"],
            "Status": x["Status"],
            "Empresa A": decimal_to_float(x["Empresa A"]),
            "Empresa B": decimal_to_float(x["Empresa B"]),
            "Motorista A": decimal_to_float(x["Motorista A"]),
            "Motorista B": decimal_to_float(x["Motorista B"]),
            "Dif Empresa": decimal_to_float(x["Dif Empresa"]),
            "Dif Motorista": decimal_to_float(x["Dif Motorista"]),
            "Maior Diferença": decimal_to_float(x["Maior Diferença"]),
            "Margem B": x["Margem B"],
            "Página A": x["Página A"],
            "Página B": x["Página B"],
        })

    return pd.DataFrame(rows)


def testar_parser_basico(caminho_atua, caminho_gw):
    resultado = auditar(caminho_atua, caminho_gw, Decimal("0.50"))

    a = resultado["registros_a"]
    b = resultado["registros_b"]

    assert a["1752"]["empresa"] == Decimal("23919.00")
    assert a["1752"]["motorista"] == Decimal("24839.65")
    assert b["1752"]["empresa"] == Decimal("23919.00")
    assert b["1752"]["motorista"] == Decimal("24839.88")

    assert a["1753"]["empresa"] == Decimal("12892.50")
    assert a["1753"]["motorista"] == Decimal("13388.62")
    assert b["1753"]["empresa"] == Decimal("12892.50")
    assert b["1753"]["motorista"] == Decimal("12892.50")

    print("Parser OK.")
    print(resultado["resumo"])
    print(resultado["debug"])

    return resultado

# ---------------------------------------------------------------------------
# Histórico & Export
# ---------------------------------------------------------------------------

HIST_PATH = Path("historico_auditoria.json")

def _normalizar_historico(hist) -> list:
    if not isinstance(hist, list):
        return []
    return [item for item in hist[:MAX_HISTORY_ENTRIES] if isinstance(item, dict)]

def _gravar_historico_atomico(hist) -> None:
    HIST_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(hist, ensure_ascii=False, indent=2)
    with tempfile.NamedTemporaryFile(
        "w",
        encoding="utf-8",
        delete=False,
        dir=HIST_PATH.parent,
        prefix=f"{HIST_PATH.stem}_",
        suffix=".tmp",
    ) as tmp:
        tmp.write(payload)
        tmp_path = Path(tmp.name)
    os.replace(tmp_path, HIST_PATH)

def salvar_historico(nome_a, nome_b, tolerancia, resumo):
    hist = []
    if HIST_PATH.exists():
        try:
            hist = _normalizar_historico(json.loads(HIST_PATH.read_text(encoding="utf-8")))
        except Exception:
            hist = []
    hist.insert(0, {
        "data_hora": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "arquivo_a": nome_a,
        "arquivo_b": nome_b,
        "tolerancia": float(tolerancia),
        **resumo,
    })
    _gravar_historico_atomico(_normalizar_historico(hist))


def carregar_historico() -> list:
    if not HIST_PATH.exists():
        return []
    try:
        return _normalizar_historico(json.loads(HIST_PATH.read_text(encoding="utf-8")))
    except Exception:
        return []

def exportar_csv(df: pd.DataFrame) -> bytes:
    buf = io.StringIO()
    df.to_csv(buf, sep=";", index=False, encoding="utf-8-sig")
    return buf.getvalue().encode("utf-8-sig")


def exportar_excel(df, resumo, nome_a, nome_b, tolerancia) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    CORES = {
        "OK": "C6EFCE", "OK por arredondamento": "FFEB9C",
        "Divergente": "FFC7CE", "Faltante no ATUA": "DDEBF7", "Faltante no GW": "FCE4D6",
    }
    buf = io.BytesIO()
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resumo"
    ws1.append(["FreteScan Pro — Resumo de Auditoria"])
    ws1.append([f"Data/Hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"])
    ws1.append([f"Arquivo A: {nome_a}  |  Arquivo B: {nome_b}"])
    ws1.append([f"Tolerância: R$ {tolerancia:.2f}"])
    ws1.append([])
    ws1.append(["Métrica", "Valor"])
    for k, v in resumo.items():
        ws1.append([k, v])

    ws2 = wb.create_sheet("Auditoria")
    ws2.append(list(df.columns))
    ws2.freeze_panes = "A2"
    for r, row in enumerate(df.itertuples(index=False), start=2):
        ws2.append([x if pd.notna(x) else "" for x in row])
        fill = PatternFill(
            start_color=CORES.get(row.Status, "FFFFFF"),
            end_color=CORES.get(row.Status, "FFFFFF"),
            fill_type="solid",
        )
        for c in range(1, len(df.columns) + 1):
            ws2.cell(row=r, column=c).fill = fill
    for col in ws2.columns:
        w = max(len(str(cell.value or "")) for cell in col)
        ws2.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 40)

    wb.save(buf)
    return buf.getvalue()


def exportar_pdf(df, resumo, nome_a, nome_b, tolerancia) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.units import cm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5 * cm, rightMargin=1.5 * cm,
                            topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    elems = []
    elems.append(Paragraph("FreteScan Pro — Relatório de Auditoria", styles["Title"]))
    elems.append(Spacer(1, 0.3 * cm))
    elems.append(Paragraph(f"Data/Hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", styles["Normal"]))
    elems.append(Paragraph(f"Arquivo A: {nome_a}  |  Arquivo B: {nome_b}", styles["Normal"]))
    elems.append(Paragraph(f"Tolerância: R$ {tolerancia:.2f}  —  Diferença = A − B", styles["Normal"]))
    elems.append(Spacer(1, 0.5 * cm))

    elems.append(Paragraph("Resumo Geral", styles["Heading2"]))
    res_data = [["Métrica", "Valor"]] + [[k, str(v)] for k, v in resumo.items()]
    t = Table(res_data, colWidths=[8 * cm, 4 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a237e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    elems.append(t)
    elems.append(Spacer(1, 0.5 * cm))

    df_div = df[df["Status"].isin(["Divergente", "Faltante no ATUA", "Faltante no GW"])].head(50)
    if not df_div.empty:
        elems.append(Paragraph("Divergências e Faltantes (até 50)", styles["Heading2"]))
        cols = ["CTE", "Status", "Empresa A", "Empresa B", "Dif. Empresa",
                "Motorista A", "Motorista B", "Dif. Motorista"]
        data = [cols] + [
            ["" if pd.isna(row[c]) else str(row[c]) for c in cols]
            for _, row in df_div.iterrows()
        ]
        cw = [2 * cm, 4 * cm, 3 * cm, 3 * cm, 3 * cm, 3 * cm, 3 * cm, 3 * cm]
        t2 = Table(data, colWidths=cw, repeatRows=1)
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a237e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
        ]))
        elems.append(t2)

    doc.build(elems)
    return buf.getvalue()
